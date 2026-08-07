from openpyxl import Workbook
from openpyxl.utils import get_column_letter


class ExcelWriter:


    def __init__(self, filename):

        self.filename = filename



    def save_jobs(self, jobs):

        print("Creating Excel file...")


        workbook = Workbook()

        sheet = workbook.active

        sheet.title = "Java Jobs"



        if not jobs:

            print("No jobs found")

            return



        # headers

        headers = jobs[0].keys()


        sheet.append(
            list(headers)
        )



        # data rows

        for job in jobs:

            sheet.append(
                list(job.values())
            )



        # auto column width

        for column in sheet.columns:

            max_length = 0

            column_letter = get_column_letter(
                column[0].column
            )


            for cell in column:

                if cell.value:

                    max_length = max(
                        max_length,
                        len(str(cell.value))
                    )


            sheet.column_dimensions[
                column_letter
            ].width = max_length + 5



        workbook.save(
            self.filename
        )


        print(
            f"Excel saved: {self.filename}"
        )